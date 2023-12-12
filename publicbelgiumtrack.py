import locale
import os

import pandas as pd
from openpyxl.reader.excel import load_workbook

from Arrest import Arrest
from webscraper import WebScraper

FILE_PATH_EXCEL = "result/2023_Marchés et travaux publics2.xlsx"

locale.setlocale(locale.LC_ALL, 'fr_BE.UTF-8')


class PublicBelgiumTrack:

    def __init__(self):
        self.pd = pd
        self.load_workbook=load_workbook

    def write_to_excel(self, arrests, file_path=FILE_PATH_EXCEL, begin_line=0):
        df = self.pd.DataFrame([arrest.as_dict() for arrest in arrests])
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name='Feuille1', index=False, startrow=begin_line, header=False)
        except FileNotFoundError:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, sheet_name='Feuille1', index=False)

    def read_from_excel(self, file_path=FILE_PATH_EXCEL):
        if os.path.isfile(file_path):
            return self.pd.read_excel(file_path)
        else:
            print(f"Le fichier n'existe pas.")


def main():
    public_belgium_track = PublicBelgiumTrack()
    previous_df = public_belgium_track.read_from_excel(FILE_PATH_EXCEL)
    last_arrest = None
    begin_line = 0
    if previous_df is not None:
        last_arrest = Arrest.from_dic(previous_df.tail(1).to_dict(orient='records')[0])
        begin_line = len(previous_df)+1
    webscraper = WebScraper()
    arrests = webscraper.extract_arrets(2023, last_arrest)
    public_belgium_track.write_to_excel(arrests, FILE_PATH_EXCEL, begin_line)


if __name__ == "__main__":
    main()
