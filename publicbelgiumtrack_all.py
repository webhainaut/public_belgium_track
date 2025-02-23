import locale
import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl.reader.excel import load_workbook

from main.Arrest import Arrest
from main.app.public_track_service import PublicTrackService

locale.setlocale(locale.LC_ALL, 'fr_BE.UTF-8')
logging.basicConfig(level=logging.INFO)


class PublicBelgiumTrack:
    YEAR = 2024
    # YEAR = datetime.now().year
    FILE_PATH_EXCEL_FORMAT = "result/CE VIe ch. {year}.xlsx"

    def __init__(self, result_path=FILE_PATH_EXCEL_FORMAT, year=YEAR):
        self.pd = pd
        self.load_workbook = load_workbook
        self.result_path = result_path.format(year=year)
        self.previous_df = None
        self.begin_line = 0
        self.last_arrest = None
        self.public_track_service = PublicTrackService()

    def write_to_excel(self, arrests):
        if arrests:
            df = self.pd.DataFrame([arrest.as_dict() for arrest in arrests])
            if not df.empty:
                df = df.sort_values(by=Arrest.REF)
            try:
                with pd.ExcelWriter(self.result_path, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
                    df.to_excel(writer, sheet_name='Feuille1', index=False, startrow=self.begin_line, header=False)
                print("remlir a partir de {begin_line} - last arrest : {arrest}".format(begin_line=self.begin_line,
                                                                                        arrest=self.last_arrest.as_dict()))
            except FileNotFoundError:
                with pd.ExcelWriter(self.result_path, engine='openpyxl', mode='w') as writer:
                    df.to_excel(writer, sheet_name='Feuille1', index=False)

    # def read_from_excel(self):
    #     if os.path.isfile(self.result_path):
    #         return self.pd.read_excel(self.result_path)
    #     else:
    #         print(f"Le fichier n'existe pas.")

    def last_line(self):
        if self.previous_df is not None:
            return len(self.previous_df) + 1
        return 0

    # def find_last_arrest(self, current_arrests):
    #     self.previous_df = self.read_from_excel()
    #     self.begin_line = self.last_line()
    #     if current_arrests:
    #         self.last_arrest = current_arrests[-1]

    # def get_current_arrests(self):
    #     self.previous_df = self.read_from_excel()
    #     current_arrests = []
    #     if self.previous_df is not None:
    #         current_arrests = [Arrest.from_dic(arrest) for arrest in self.previous_df.to_dict(orient='records')]
    #     return current_arrests

    def find_last_arrest_from_previous_df(self):
        if self.previous_df is not None:
            return Arrest.from_dic(self.previous_df.tail(1).to_dict(orient='records')[0])



def main():
    public_belgium_track = PublicBelgiumTrack()

    # public_belgium_track.public_track_service.download(259209)
    public_belgium_track.public_track_service.download_all(REFS)
    # public_belgium_track.public_track_service.update_year(2024)
    # public_belgium_track.public_track_service.update_all(REFS)
    # public_belgium_track.public_track_service.update(260454)


    # current_arrests = public_belgium_track.get_current_arrests()
    # public_belgium_track.find_last_arrest(current_arrests)
    # arrests = public_belgium_track.get_arrests(REFS)
    # public_belgium_track.write_to_excel(arrests)


# REFS = [260454, 260616]
# REFS = [247478, 255267, 255438, 255470, 255472, 255668, 255678, 255679, 255681, 255844, 255962, 255964, 256014, 256484,
#         256672, 257478, 257654, 257819, 257919, 257921, 258204, 255824, 258274, 255568, 258317, 256090, 256352, 256552,
#         256835, 256952, 257003, 255570, 257985, 257647, 257248, 257984, 258070, 257009, 258245, 256675, 261996, 258994,
#         259019, 259068, 259209]
REFS = [259019, 259068, 259209]
if __name__ == "__main__":
    start_time = datetime.now()
    main()
    end_time = datetime.now()
    execution_time = end_time - start_time
    print(f"Le script a pris {execution_time} secondes pour s'exécuter.")
